
import os
import re
import pandas as pd

import openpyxl
import numpy as np
from openpyxl import load_workbook
import glob

# import pathlib
# p = pathlib.Path('./')
# for txt_file in p.glob(pattern='*.txt'):
#     print (txt_file)


# 查找所有后缀为 .txt 的文本文件
txt_files = glob.glob('*.txt')
lines=[]
txt_file_names=[]
#print(txt_files)
for txt_file in txt_files:
# 读取文本文件并按行分割
    with open(txt_file, 'r',encoding="utf-8") as file:
        this_lines = file.readlines()
        lines=lines+this_lines
        txt_file_names.append(txt_file)

#print(lines)


# 定义正则表达式模式
case_pattern = re.compile(r'ispd24contest/(.+)\.def')

gcell_pattern = re.compile(r'gcell grid:\s+(\d+ x \d+ x \d+)')
#unit_length_wire_pattern = re.compile(r'unit length wire:\s+(\d+\.\d+)')
#unit_via_pattern = re.compile(r'unit via:\s+(\d+)')
#num_nets_pattern = re.compile(r'num of nets :\s+(\d+)')
time_pattern = re.compile(r'\[\s*([\d.]+)\]')
#step_routed_pattern = re.compile(r'step routed #nets:\s+(\d+),\s+(\d+),\s+(\d+)')
#stage1_routed_net_num_pattern = re.compile(r'num of nets :        \s+([\d.]+)')
#three_stage_routed_net_num_pattern = re.compile(r'\[\s+([\d.]+)\] \s+([\d.]+) / \s+([\d.]+) nets have overflows.')#有几个stage就会匹配几次
three_stage_routed_net_num_pattern = re.compile(r'\[\s*([\d.]+)\]\s+(\d+)\s/\s+(\d+)\snets have overflows')#有几个stage就会匹配几次

#step_time_pattern = re.compile(r'step time consumption:\s+([\d.]+) s,\s+([\d.]+) s,\s+([\d.]+) s')
#cost_pattern = re.compile(r'(wire|via|overflow|total) cost:\s+([\d.]+)')
total_time_pattern = re.compile(r'\[([\d.]+)\] Terminated.')
wire_length_metric_pattern = re.compile(r'wire length \(metric\):\s+([\d.]+)')
total_via_count_pattern = re.compile(r'total via count:\s+([\d.]+)')
total_wire_overflow_pattern = re.compile(r'total wire overflow:\s+([\d.]+)')
#num_of_open_nets_pattern = re.compile(r'use evaluator test data: Number of open nets :\s(\d+)')
#evaluator_wirelength_cost_pattern= re.compile(r'use evaluator test data: wirelength cost\s+([\d.]+)')
#evaluator_via_cost_pattern= re.compile(r'use evaluator test data: via cost\s+([\d.]+)')
#evaluator_overflow_cost_pattern= re.compile(r'use evaluator test data: overflow cost\s+([\d.]+)')
#evaluator_total_cost_pattern= re.compile(r'use evaluator test data: total cost\s+([\d.]+)')



timestamp_keyword={

}


# 初始化数据存储字典
data = {
    'Environment': [],
    'Case_Name': [],
    #'scaled_score':[],
    #'runtime_factor':[],
    #'evaluator_total_cost': [],
    #'Total_Cost': [],
    'Total_Time': [],

    'Stage2_Routed_Nets_Num': [],
    'Stage3_Routed_Nets_Num': [],
    'finally_overflow_nets_num': [],
    'wire length(metric)':[],
    'total via count':[],
    'total wire overflow':[],


    #'num of open nets': [],
    'GCell_Grid': [],
    #'Unit_Length_Wire': [],
    #'Unit_Via': [],
    'Num_of_Nets': [],
    #'Total_Time': [],
    'Parser_Time': [],
    'Parser_Time_Proportion': [],
    'Stage1_Routed_Nets_Num': [],
    'Stage1_Routed_Time': [],
    'Stage1_Routed_Time_Proportion': [],

    'generate_all_two_pin_net_time':[],
    'generate_all_two_pin_net_time_proportion':[],
    'generate_congestion_map':[],
    'generate_congestion_map_proportion':[],
    'generate_treeforest':[],
    'generate_treeforest_proportion':[],
    'generate_link_P_tree_to_P_pattern_mask':[],
    'generate_link_P_tree_to_P_pattern_mask_proportion':[],
    'generate_capacity_map':[],
    'generate_capacity_map_proportion':[],
    'generate_net_pass_through_gcell_mask':[],
    'generate_net_pass_through_gcell_mask_proportion':[],
    'optimize_time':[],
    'optimize_time_proportion':[],
   


    
    'Stage2_Routed_Time': [],
    'Stage2_Routed_Time_Proportion': [],
    
    'Stage3_Routed_Time': [],
    'Stage3_Routed_Time_Proportion': [],
    'other_time': [],
    'other_time_proportion': [],
    #'evaluator_wirelength_cost': [],
    #'Wire_Cost': [],
    #'Wire_Cost_Proportion': [],
    #'evaluator_via_cost': [],
    #'Via_Cost': [],
    #'Via_Cost_Proportion': [],
    #'Overflow_Cost': [],
    #'evaluator_overflow_cost': [],
    #'Overflow_Cost_Proportion': []
    
    
}

# 逐行解析文本内容
case_name = ''
total_time = 0.0
parser_time = 0.0
wire_cost = 0.0
via_cost = 0.0
overflow_cost = 0.0

stage_count=1

#temp_pointer_timestamp=''

for line in lines:
    case_match = re.search(case_pattern, line)
    #print(case_match)
    gcell_match = re.search(gcell_pattern, line)
    #unit_length_wire_match = re.search(unit_length_wire_pattern, line)
    #unit_via_match = re.search(unit_via_pattern, line)
    #num_nets_match = re.search(num_nets_pattern, line)
    time_match = re.search(time_pattern, line)
    three_stage_routed_net_num_match = re.search(three_stage_routed_net_num_pattern, line)
    # print(three_stage_routed_net_num_match)
    #print(three_stage_routed_net_num_pattern)
    #step_routed_match = re.search(step_routed_pattern, line)
    #print(step_routed_match)
    #step_time_match = re.search(step_time_pattern, line)
    total_time_match = re.search(total_time_pattern, line)
    #cost_match = re.search(cost_pattern, line)
    wire_length_metric_match = re.search(wire_length_metric_pattern, line)
    total_via_count_match = re.search(total_via_count_pattern, line)
    total_wire_overflow_match = re.search(total_wire_overflow_pattern, line)

    #num_of_open_nets_match = re.search(num_of_open_nets_pattern, line)
    #evaluator_total_cost_match = re.search(evaluator_total_cost_pattern, line)
    #evaluator_wirelength_cost_match = re.search(evaluator_wirelength_cost_pattern, line)
    #evaluator_via_cost_match = re.search(evaluator_via_cost_pattern, line)
    #evaluator_overflow_cost_match = re.search(evaluator_overflow_cost_pattern, line)
    

    # if num_of_open_nets_match:
    #     data['num of open nets'].append(num_of_open_nets_match.group(1))
    # if evaluator_total_cost_match:
    #     data['evaluator_total_cost'].append(evaluator_total_cost_match.group(1))
    #     data['Wire_Cost_Proportion'].append(round(float(data['evaluator_wirelength_cost'][-1]) /float( data['evaluator_total_cost'][-1]) * 100, 2))
    #     data['Via_Cost_Proportion'].append(round(float(data['evaluator_via_cost'][-1]) /float( data['evaluator_total_cost'][-1]) * 100, 2))
    #     data['Overflow_Cost_Proportion'].append(round(float(data['evaluator_overflow_cost'][-1]) /float( data['evaluator_total_cost'][-1]) * 100, 2))
    # if evaluator_wirelength_cost_match:
    #     data['evaluator_wirelength_cost'].append(evaluator_wirelength_cost_match.group(1))
    # if evaluator_via_cost_match:
    #     data['evaluator_via_cost'].append(evaluator_via_cost_match.group(1))
    # if evaluator_overflow_cost_match:
    #     data['evaluator_overflow_cost'].append(evaluator_overflow_cost_match.group(1))

    if case_match:
        #data['Environment'].append(txt_file)
        #data['Environment'].append(txt_file[ :txt_file.index('.')])
        case_name = case_match.group(1).split('/')[-1].split('.')[0]
        data['Case_Name'].append(case_name)
        #print(case_name)
    if gcell_match:
        data['GCell_Grid'].append(gcell_match.group(1))
        #print(gcell_match.group(1))
    # if unit_length_wire_match:
    #     data['Unit_Length_Wire'].append(unit_length_wire_match.group(1))
        #print(unit_length_wire_match.group(1))
    # if unit_via_match:
    #     data['Unit_Via'].append(unit_via_match.group(1))
    #     #print(unit_via_match.group(1))
    # if num_nets_match:
    #     data['Num_of_Nets'].append(num_nets_match.group(1))
    #     #print(num_nets_match.group(1))
    if time_match:
        timestamp = float(time_match.group(1))
        #total_time = max(total_time, timestamp)
        #print(f"Total time: {total_time} seconds")
        # if 'total cost(ispd24 score)' in line:
        #     total_time = timestamp
        #     data['Total_Time'].append(total_time)
        #     #print(f"Total time: {total_time} seconds")
        
        if 'Finished reading lef/def' in line:
            parser_time = timestamp
            data['Parser_Time'].append(parser_time)
            temp_pointer_timestamp='parser_time'
            #print(f"Parser time: {parser_time} seconds")
        #else:data['Parser_Time'].append(0)
        if 'generate_all_two_pin_net_finish' in line:
            generate_all_two_pin_net_finish_time = timestamp
            data['generate_all_two_pin_net_time'].append(generate_all_two_pin_net_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='generate_all_two_pin_net_finish_time'
        #else:data['generate_all_two_pin_net_time'].append(0)
        #if 'generate_2d_estimated_congestion_map_by_RUDY' in line:
        if 'generate_2d_estimated_congestion_map_by_steinertree' in line:
            generate_2d_estimated_congestion_map_by_RUDY_finish_time = timestamp
            data['generate_congestion_map'].append(generate_2d_estimated_congestion_map_by_RUDY_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='generate_2d_estimated_congestion_map_by_RUDY_finish_time'
        #else:data['generate_congestion_map'].append(0)
        if 'generate_treeforest_finish' in line:
            generate_treeforest_finish_time = timestamp
            data['generate_treeforest'].append(generate_treeforest_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='generate_treeforest_finish_time'
        #else:data['generate_treeforest'].append(0)
        if 'link_P_tree_to_P_pattern_mask_finish' in line:
            link_P_tree_to_P_pattern_mask_finish_time = timestamp
            data['generate_link_P_tree_to_P_pattern_mask'].append(link_P_tree_to_P_pattern_mask_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='link_P_tree_to_P_pattern_mask_finish_time'
        #else:data['generate_link_P_tree_to_P_pattern_mask'].append(0)
        if '2d_capacitymap_finish' in line:
            generate_2d_capacitymap_finish_time = timestamp
            data['generate_capacity_map'].append(generate_2d_capacitymap_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='generate_2d_capacitymap_finish_time'
        #else:data['generate_capacity_map'].append(0)
        if 'generate_net_pass_through_gcell_mask_finish' in line:
            generate_net_pass_through_gcell_mask_finish_time = timestamp
            data['generate_net_pass_through_gcell_mask'].append(generate_net_pass_through_gcell_mask_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='generate_net_pass_through_gcell_mask_finish_time'
        #else:data['generate_net_pass_through_gcell_mask'].append(0)
        if 'Torch optimization stopped' in line:
            torch_optimization_stopped_time = timestamp
            data['optimize_time'].append(torch_optimization_stopped_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='torch_optimization_stopped_time'
        #else:data['optimize_time'].append(0)
        if 'stage 2: pattern routing with possible detours' in line:
            stage1_finish_time = timestamp
            data['Stage1_Routed_Time'].append(stage1_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='stage1_finish_time'
        #else:data['Stage1_Routed_Time'].append(0)
        if 'stage 3: maze routing on sparsified routing graph' in line:
            stage2_finish_time = timestamp
            data['Stage2_Routed_Time'].append(stage2_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='stage2_finish_time'
        #else:data['Stage2_Routed_Time'].append(0)
        if 'routing statistics' in line:
            stage3_finish_time = timestamp
            data['Stage3_Routed_Time'].append(stage3_finish_time - eval(temp_pointer_timestamp))
            temp_pointer_timestamp='stage3_finish_time'
        #else:data['Stage3_Routed_Time'].append(0)
        

        

        if 'Terminated.' in line:
            total_time = timestamp
            data['Total_Time'].append(total_time)
            data['other_time'].append(total_time - stage3_finish_time)


            data['Parser_Time_Proportion'].append(round(data['Parser_Time'][-1] / total_time,2))
            data['Stage1_Routed_Time_Proportion'].append(round(data['Stage1_Routed_Time'][-1] / total_time,2))
            data['Stage2_Routed_Time_Proportion'].append(round(data['Stage2_Routed_Time'][-1] / total_time,2))
            data['Stage3_Routed_Time_Proportion'].append(round(data['Stage3_Routed_Time'][-1] / total_time,2))
            data['other_time_proportion'].append(((total_time - stage3_finish_time) / total_time) * 100)


            temp_pointer_terminated='parser_time'#指向目前能找到的变量的名称
            try:#下面这一部分代码是为了兼容，因为我想兼容原版cugr2的log内容，并且兼容一些旧版本的log内容（主要目的）
                data['generate_all_two_pin_net_time_proportion'].append(round((generate_all_two_pin_net_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='generate_all_two_pin_net_finish_time'
            except NameError:
                data['generate_all_two_pin_net_time_proportion'].append('_')

            try:
                data['generate_congestion_map_proportion'].append(round((generate_2d_estimated_congestion_map_by_RUDY_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='generate_2d_estimated_congestion_map_by_RUDY_finish_time'
            except NameError:
                data['generate_congestion_map_proportion'].append('_')
            
            try:
                data['generate_treeforest_proportion'].append(round((generate_treeforest_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='generate_treeforest_finish_time'
            except NameError:
                data['generate_treeforest_proportion'].append('_')
            
            try:
                data['generate_link_P_tree_to_P_pattern_mask_proportion'].append(round((link_P_tree_to_P_pattern_mask_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='link_P_tree_to_P_pattern_mask_finish_time'
            except NameError:
                data['generate_link_P_tree_to_P_pattern_mask_proportion'].append('_')
                
            try:
                data['generate_capacity_map_proportion'].append(round((generate_2d_capacitymap_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='generate_2d_capacitymap_finish_time'
            except NameError:
                data['generate_capacity_map_proportion'].append('_')

            try:
                data['generate_net_pass_through_gcell_mask_proportion'].append(round((generate_net_pass_through_gcell_mask_finish_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='generate_net_pass_through_gcell_mask_finish_time'
            except NameError:
                data['generate_net_pass_through_gcell_mask_proportion'].append('_')
                
            try:
                data['optimize_time_proportion'].append(round((torch_optimization_stopped_time - eval(temp_pointer_terminated))/total_time,2))
                temp_pointer_terminated='torch_optimization_stopped_time'
            except NameError:
                data['optimize_time_proportion'].append('_')
            # data['generate_all_two_pin_net_time_proportion'].append(round((generate_all_two_pin_net_finish_time - parser_time)/total_time,2))
            # data['generate_congestion_map_proportion'].append(round((generate_2d_estimated_congestion_map_by_RUDY_finish_time - generate_all_two_pin_net_finish_time)/total_time,2))
            # data['generate_treeforest_proportion'].append(round((generate_treeforest_finish_time - generate_2d_estimated_congestion_map_by_RUDY_finish_time)/total_time,2))
            # data['generate_link_P_tree_to_P_pattern_mask_proportion'].append(round((link_P_tree_to_P_pattern_mask_finish_time - generate_treeforest_finish_time)/total_time,2))
            # data['generate_capacity_map_proportion'].append(round((generate_2d_capacitymap_finish_time - link_P_tree_to_P_pattern_mask_finish_time)/total_time,2))
            # data['generate_net_pass_through_gcell_mask_proportion'].append(round((generate_net_pass_through_gcell_mask_finish_time - generate_2d_capacitymap_finish_time)/total_time,2))
            # data['optimize_time_proportion'].append(round((torch_optimization_stopped_time - generate_net_pass_through_gcell_mask_finish_time)/total_time,2))
            stage_count=1#重新计数，以免有的小case下不具有三个stage
            

    if three_stage_routed_net_num_match:
        #timestamp=float(three_stage_routed_net_num_match.group(1))

        if (stage_count%3==1):
            data['Stage1_Routed_Nets_Num'].append(three_stage_routed_net_num_match.group(3))
            data['Num_of_Nets'].append(three_stage_routed_net_num_match.group(3))
            data['Stage2_Routed_Nets_Num'].append(three_stage_routed_net_num_match.group(2))
        if (stage_count%3==2):
            data['Stage3_Routed_Nets_Num'].append(three_stage_routed_net_num_match.group(2))
        if (stage_count%3==0):
            Stage3_Routed_Nets_Num=three_stage_routed_net_num_match.group(2)
            data['finally_overflow_nets_num'].append(three_stage_routed_net_num_match.group(2))
        stage_count=stage_count+1

    if wire_length_metric_match:
        data['wire length(metric)'].append(wire_length_metric_match.group(1))
    if total_via_count_match:
        data['total via count'].append(total_via_count_match.group(1))
    if total_wire_overflow_match:
        data['total wire overflow'].append(total_wire_overflow_match.group(1))
            #print(f"Parser time proportion: {data['Parser_Time_Proportion'][-1]}")
    # if step_routed_match:
    #     data['Stage1_Routed_Nets_Num'].append(step_routed_match.group(1))
    #     #print(step_routed_match.group(1))
    #     data['Stage2_Routed_Nets_Num'].append(step_routed_match.group(2))
    #     #print(step_routed_match.group(2))
    #     data['Stage3_Routed_Nets_Num'].append(step_routed_match.group(3))
    #     #print(step_routed_match.group(3))

    # if step_time_match:
    #     data['Stage1_Routed_Time'].append(step_time_match.group(1))
        
        
        
    #     #print(step_time_match.group(1))
    #     data['Stage2_Routed_Time'].append(step_time_match.group(2))
    #     #print(step_time_match.group(2))
    #     data['Stage3_Routed_Time'].append(step_time_match.group(3))
    #     #print(step_time_match.group(3))

    #     not_precise_total_time = float(step_time_match.group(1))+float(step_time_match.group(2))+float(step_time_match.group(3))+parser_time
    #     #data['Total_Time'].append(not_precise_total_time)
    #     #print(precise_total_time)
    #     data['Stage1_Routed_Time_Proportion'].append(round(float(step_time_match.group(1)) / not_precise_total_time, 2))
    #     #print(f"Stage1 time proportion: {data['Stage1_Routed_Time_Proportion'][-1]}")
    #     data['Stage2_Routed_Time_Proportion'].append(round(float(step_time_match.group(2)) / not_precise_total_time, 2))
    #     #print(f"Stage2 time proportion: {data['Stage2_Routed_Time_Proportion'][-1]}")
    #     data['Stage3_Routed_Time_Proportion'].append(round(float(step_time_match.group(3)) / not_precise_total_time, 2))
    #     #print(f"Stage3 time proportion: {data['Stage3_Routed_Time_Proportion'][-1]}")
    #     data['Parser_Time_Proportion'].append(round(parser_time / not_precise_total_time, 2))
    if total_time_match:
        total_time = float(total_time_match.group(1))
        #print(f"Total time: {total_time} seconds")
        #data['Total_Time'].append(total_time)
        #data['other_time'].append(total_time - not_precise_total_time)
        #print(f"Total time: {total_time} seconds")
        
    # if cost_match:
    #     cost_type = cost_match.group(1)
    #     #print(cost_match.group(1))
    #     #print(cost_type)
    #     cost_value = float(cost_match.group(2))
    #     if cost_type == 'wire':
    #         wire_cost = cost_value
    #         #print(f"Wire cost: {wire_cost}")
    #         data['Wire_Cost'].append(wire_cost)
    #     elif cost_type == 'via':
    #         via_cost = cost_value
    #         #print(f"Via cost: {via_cost}")
    #         data['Via_Cost'].append(via_cost)
    #     elif cost_type == 'overflow':
    #         overflow_cost = cost_value
    #         #print(f"Overflow cost: {overflow_cost}")
    #         data['Overflow_Cost'].append(overflow_cost)
    #         total_cost = wire_cost + via_cost + overflow_cost
    #         #print(f"Total cost: {total_cost}")
    #         data['Total_Cost'].append(total_cost)
    #         data['Wire_Cost_Proportion'].append(round((wire_cost / total_cost) * 100, 2))
    #         data['Via_Cost_Proportion'].append(round((via_cost / total_cost) * 100, 2))
    #         data['Overflow_Cost_Proportion'].append(round((overflow_cost / total_cost) * 100, 2))
print(data)
# if len(data['num of open nets'])==0:
#     data['num of open nets'] = [0]*len(data['Total_Cost'])
#     data['evaluator_total_cost'] = data['Total_Cost']
#     data['evaluator_wirelength_cost'] = data['Wire_Cost']
#     data['evaluator_via_cost'] = data['Via_Cost']
#     data['evaluator_overflow_cost'] = data['Overflow_Cost']
            
# median_time = []
# #以cpu环境下的数据做
# for txt_file_name in txt_file_names:
#     for case_index in range(int(len(data['Case_Name'])/len(txt_file_names))):
#         data['Environment'].append(txt_file_name[ :txt_file_name.index('.')])
#         if data['Environment'][-1] == 'cpu':
#             median_time.append(data['Total_Time'][len(data['Environment'])-1])

# #print(median_time)

# for txt_file_name in txt_file_names:
#     for case_index in range(int(len(data['Case_Name'])/len(txt_file_names))):
#         runtime_factor=[0.02*np.log2(float(data['Total_Time'][len(data['score'])])/float(median_time[case_index])),0.2,-0.2]
#         #print(runtime_factor)
#         data['runtime_factor'].append(str(np.median(runtime_factor)))
#         data['score'].append(str(float(data['evaluator_total_cost'][len(data['score'])])*(np.median(runtime_factor)+1)))



#从官网找到的中位数数据
# median_time = {
#     "ariane133_51":14.868,
#     "ariane133_68":16.268,
#     "bsg_chip":120.810,
#     "mempool_tile":16.968,
#     "nvdla":44.730,
#     "mempool_group":2974.426,
#     "mempool_cluster":2974.426,
#     "mempool_cluster_large":2974.426
# }

# best_team_raw_score={
#     "ariane133_51":22765372,
#     "ariane133_68":20251238,
#     "bsg_chip":113537664,
#     "mempool_tile":15402753,
#     "nvdla":49909986,
#     "mempool_group":409309772,
#     "mempool_cluster":1663344708,
#     "mempool_cluster_large":20438460750
# }

# best_team_scaled_score={
#     "ariane133_51":22100882,
#     "ariane133_68":20014313,
#     "bsg_chip":110393434,
#     "mempool_tile":14867672,
#     "nvdla":47592238,
#     "mempool_group":391210938,
#     "mempool_cluster":1593513066,
#     "mempool_cluster_large":20402113329
# }
cugr2_original_runtime={
    "ariane133_51":14.868,
    "ariane133_68":16.268,
    "bsg_chip":120.810,
    "mempool_tile":16.967,
    "nvdla":44.730,
    "mempool_group":2974.426,
    "cluster":47192.669,
    "mempool_cluster_large":2974.426
}

cugr2_stage2_routed_num={
    "ariane133_51":1532,
    "ariane133_68":3682,
    "bsg_chip":14977,
    "mempool_tile":2499,
    "nvdla":10780,
    "mempool_group":448792,
    "cluster":1937168,
    "mempool_cluster_large":0.0
}
cugr2_stage3_routed_num={
    "ariane133_51":33,
    "ariane133_68":99,
    "bsg_chip":1349,
    "mempool_tile":194,
    "nvdla":487,
    "mempool_group":172575,
    "cluster":1051045,
    "mempool_cluster_large":0.0
}
cugr2_after_stage3_overflow_net_num={
    "ariane133_51":0,
    "ariane133_68":1,
    "bsg_chip":229,
    "mempool_tile":22,
    "nvdla":36,
    "mempool_group":237242,
    "cluster":1374676,
    "mempool_cluster_large":0.0
}
cugr2_wire_length={
    "ariane133_51":18652433,
    "ariane133_68":19046238,
    "bsg_chip":118017552,
    "mempool_tile":16999672,
    "nvdla":43681146,
    "mempool_group":599479848,
    "cluster":2520387756,
    "mempool_cluster_large":0.0
}
cugr2_total_via_count={
    "ariane133_51":1008862,
    "ariane133_68":964035,
    "bsg_chip":6357593,
    "mempool_tile":1144713,
    "nvdla":1464784,
    "mempool_group":26619488,
    "cluster":99471802,
    "mempool_cluster_large":0
}
cugr2_total_wire_overflow={
    "ariane133_51":0,
    "ariane133_68":1,
    "bsg_chip":85,
    "mempool_tile":3,
    "nvdla":15,
    "mempool_group":183732,
    "cluster":1045424,
    "mempool_cluster_large":0.0
}
for txt_file_name in txt_file_names:
    for case_index in range(int(len(data['Case_Name'])/len(txt_file_names))):
        data['Environment'].append(txt_file_name[ :txt_file_name.index('.')])
        # case_name = data['Case_Name'][len(data['scaled_score'])]
        # runtime_factor=[0.02*np.log2(float(data['Total_Time'][len(data['scaled_score'])])/float(median_time[f'{case_name}'])),0.2,-0.2]
        # #print(runtime_factor)
        # data['runtime_factor'].append(str(np.median(runtime_factor)))
        # #data['scaled_score'].append(str(float(data['evaluator_total_cost'][len(data['scaled_score'])-1])*(np.median(runtime_factor)+1)))
        # data['scaled_score'].append(str(float(data['evaluator_total_cost'][len(data['scaled_score'])])*(np.median(runtime_factor)+1)))



for obj in data:#对于缺数据的情况，进行补齐
    if data[obj] == []:
        for i in range(len(data['Case_Name'])):
            data[obj].append('no data')
        print(f'{obj} is empty')



print(data)

# 创建 DataFrame
df = pd.DataFrame(data)


if os.path.exists(os.path.join("",'output.xlsx')):#如果根目录下已经有了output.xlsx，则追加数据
    # 写入 Excel 文件
    with pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a') as writer:
        # 加载现有的 Excel 文件
        writer.workbook = load_workbook('output.xlsx')
        #print('writer.workbook')

        # 将数据写入现有的工作表
        for case in df['Case_Name'].unique():
            case_df = df[df['Case_Name'] == case]
            sheet_name = case

            # 如果工作表已存在，则追加数据
            if sheet_name in writer.workbook.sheetnames:
                
                startrow = writer.workbook[sheet_name].max_row
                for row in pd.DataFrame(case_df).iterrows():
                    values = row[1].to_list()
                    
                    for col in range(len(values)):
                        writer.workbook[sheet_name].cell(row=startrow + 1, column=col + 1, value=values[col])
                    
                    
            else:
                case_df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.workbook.save('output.xlsx')
    writer.workbook.close() 
else:#如果根目录下没有output.xlsx


    # 写入 Excel 文件
    with pd.ExcelWriter('output.xlsx') as writer:
        for case in df['Case_Name'].unique():
            case_df = df[df['Case_Name'] == case]
            case_df.to_excel(writer, sheet_name=case, index=False)
    #writer.workbook.save('output.xlsx')
    #writer.workbook.close()








from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 读取xlsx文件
file_path = 'output.xlsx'
workbook = load_workbook(file_path)

# 创建一个新的工作表用于拼接数据
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = 'Combined Data'

# 指定要处理的列
#columns_to_process = [2,3,4,5,6,24,27,31]

columns_to_process = [list(data).index('Total_Time')+1,list(data).index('Stage2_Routed_Nets_Num')+1,list(data).index('Stage3_Routed_Nets_Num')+1,list(data).index('wire length(metric)')+1,list(data).index('total via count')+1,list(data).index('total wire overflow')+1,list(data).index('finally_overflow_nets_num')+1]

# 循环每个sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # 处理指定的列
    for col in columns_to_process:
        column_data = [sheet.cell(row=row, column=col).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None]
        if column_data:
            min_value = min(column_data)
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value == min_value:
                    sheet.cell(row=row, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")#最优结果格子底色变为红色

        # if col==list(data).index('scaled_score')+1:
        #     for row in range(2, sheet.max_row + 1):
                
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         scaled_score_temp=sheet.cell(row=row,column=list(data).index('scaled_score')+1).value
                
        #         # print(case_name_temp)
        #         # print(best_team_scaled_score[f'{case_name_temp}'])
        #         # print(sheet.cell(row=row, column=list(data).index('scaled_score')+1).value)
        #         if float(sheet.cell(row=row, column=list(data).index('scaled_score')+1).value) <= float(best_team_scaled_score[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index('scaled_score')+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         proportion_temp=float(sheet.cell(row=row, column=list(data).index('scaled_score')+1).value)/float(best_team_scaled_score[f'{case_name_temp}'])
        #         sheet.cell(row=row, column=list(data).index('scaled_score')+1).value=f'{scaled_score_temp}({proportion_temp*100:.2f})'
                
        # if col==list(data).index('evaluator_total_cost')+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         raw_score_temp=sheet.cell(row=row,column=list(data).index('evaluator_total_cost')+1).value
        #         if float(sheet.cell(row=row, column=list(data).index('evaluator_total_cost')+1).value) <= float(best_team_raw_score[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index('evaluator_total_cost')+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         proportion_temp=float(sheet.cell(row=row, column=list(data).index('evaluator_total_cost')+1).value)/float(best_team_raw_score[f'{case_name_temp}'])
        #         sheet.cell(row=row, column=list(data).index('evaluator_total_cost')+1).value=f'{raw_score_temp}({proportion_temp*100:.2f})'
        compare_objs={
            'Total_Time':'cugr2_original_runtime',
            'Stage2_Routed_Nets_Num':'cugr2_stage2_routed_num',
            'Stage3_Routed_Nets_Num':'cugr2_stage3_routed_num',
            'wire length(metric)':'cugr2_wire_length',
            'total via count':'cugr2_total_via_count',
            'total wire overflow':'cugr2_total_wire_overflow',
            'finally_overflow_nets_num':'cugr2_after_stage3_overflow_net_num',
            #'':'',
        }
        for compare_obj in compare_objs:
            if col==list(data).index(compare_obj)+1:
                for row in range(2, sheet.max_row + 1):
                    case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
                    this_data_temp=sheet.cell(row=row,column=list(data).index(compare_obj)+1).value
                    if float(sheet.cell(row=row, column=list(data).index(compare_obj)+1).value) <= float(eval(compare_objs[compare_obj])[f'{case_name_temp}']):
                        sheet.cell(row=row, column=list(data).index(compare_obj)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
                    if float(eval(compare_objs[compare_obj])[f'{case_name_temp}'])==0:
                        add_str="yuanban is 0"
                        sheet.cell(row=row, column=list(data).index(compare_obj)+1).value=f'{this_data_temp}({add_str})'
                    else:
                        add_str=float(sheet.cell(row=row, column=list(data).index(compare_obj)+1).value)/float(eval(compare_objs[compare_obj])[f'{case_name_temp}'])
                        sheet.cell(row=row, column=list(data).index(compare_obj)+1).value=f'{this_data_temp}({add_str*100:.2f}%)'
        # temp_str='Stage2_Routed_Nets_Num'
        # if col==list(data).index(temp_str)+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         this_data_temp=sheet.cell(row=row,column=list(data).index(temp_str)+1).value
        #         if float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value) <= float(cugr2_stage2_routed_num[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         if float(cugr2_stage2_routed_num[f'{case_name_temp}'])==0:
        #             add_str="yuanban is 0"
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str})'
        #         else:
        #             add_str=float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value)/float(cugr2_stage2_routed_num[f'{case_name_temp}'])
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str*100:.2f})'
        # temp_str='Stage3_Routed_Nets_Num'
        # if col==list(data).index(temp_str)+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         this_data_temp=sheet.cell(row=row,column=list(data).index(temp_str)+1).value
        #         if float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value) <= float(cugr2_stage3_routed_num[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         if float(cugr2_stage3_routed_num[f'{case_name_temp}'])==0:
        #             add_str="yuanban is 0"
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str})'
        #         else:
        #             add_str=float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value)/float(cugr2_stage3_routed_num[f'{case_name_temp}'])
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str*100:.2f})'
        # temp_str='wire length(metric)'
        # if col==list(data).index(temp_str)+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         this_data_temp=sheet.cell(row=row,column=list(data).index(temp_str)+1).value
        #         if float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value) <= float(cugr2_wire_length[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         if float(cugr2_wire_length[f'{case_name_temp}'])==0:
        #             add_str="yuanban is 0"
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str})'
        #         else:
        #             add_str=float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value)/float(cugr2_wire_length[f'{case_name_temp}'])
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str*100:.2f})'
        # temp_str='total via count'
        # if col==list(data).index(temp_str)+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         this_data_temp=sheet.cell(row=row,column=list(data).index(temp_str)+1).value
        #         if float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value) <= float(cugr2_total_via_count[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         if float(cugr2_total_via_count[f'{case_name_temp}'])==0:
        #             add_str="yuanban is 0"
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str})'
        #         else:
        #             add_str=float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value)/float(cugr2_total_via_count[f'{case_name_temp}'])
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str*100:.2f})'
        # temp_str='total wire overflow'
        # if col==list(data).index(temp_str)+1:
        #     for row in range(2, sheet.max_row + 1):
        #         case_name_temp=sheet.cell(row=row,column=list(data).index('Case_Name')+1).value
        #         this_data_temp=sheet.cell(row=row,column=list(data).index(temp_str)+1).value
        #         if float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value) <= float(cugr2_total_wire_overflow[f'{case_name_temp}']):
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).font = openpyxl.styles.Font(bold=True, color="00FF00")
        #         if float(cugr2_total_wire_overflow[f'{case_name_temp}'])==0:
        #             add_str="yuanban is 0"
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str})'
        #         else:
        #             add_str=float(sheet.cell(row=row, column=list(data).index(temp_str)+1).value)/float(cugr2_total_wire_overflow[f'{case_name_temp}'])
        #             sheet.cell(row=row, column=list(data).index(temp_str)+1).value=f'{this_data_temp}({add_str*100:.2f})'
        

    # 将每个sheet的内容拼接到新的工作表中
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        
        new_sheet.append([cell.value for cell in row])
        for cell in row:
            if cell.fill == PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"):#如果发现这个格子颜色是红色，也就是说这个格子的数据是最优的数据，然后在新的表格里面把这一项既标红，字体颜色也进行突出显示
                #print(cell.fill)
                new_sheet.cell(row=new_sheet.max_row, column=cell.column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                new_sheet.cell(row=new_sheet.max_row, column=cell.column).font = openpyxl.styles.Font(bold=True, color="00003366")
                #改变字体颜色
            if cell.font == openpyxl.styles.Font(bold=True, color="00FF00"):
                new_sheet.cell(row=new_sheet.max_row, column=cell.column).font = openpyxl.styles.Font(bold=True, color="00FF00")
        
    new_sheet.append([])
    # 调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 调整列宽
    for column in new_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        new_sheet.column_dimensions[column_letter].width = adjusted_width

# # 复制格式
# for row in new_sheet.iter_rows():
#     for cell in row:
#         if cell.fill is not None:
#             new_sheet.cell(row=cell.row, column=cell.column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#             #cell.fill = sheet.cell(cell.row, cell.column).fill

# 保存修改后的文件和新的工作表
workbook.save('经过数据标注的output.xlsx')
workbook.close
new_workbook.save('与原版cugr2的对比.xlsx')














# # 读取原始 Excel 文件
# file_path = 'output.xlsx'
# xl = pd.ExcelFile(file_path)

# # 读取每个 sheet 的数据到 DataFrame
# dfs = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}

# # 创建一个新的 DataFrame 用于存放所有 sheet 的数据
# combined_df = pd.DataFrame()

# # 将所有 sheet 的数据合并到一个 DataFrame 中，并在每个 DataFrame 之间插入一行空行
# for sheet_name, df in dfs.items():
#     if not combined_df.empty:
#         # 插入空行
#         empty_row = pd.Series([None] * len(combined_df.columns), index=combined_df.columns)
#         combined_df = pd.concat([combined_df, pd.DataFrame({col: [None] * 1})], ignore_index=True)
#     combined_df = pd.concat([combined_df, df], ignore_index=True)

# # 创建一个新的 Excel 文件并将合并后的数据写入新的 sheet
# output_file_path = 'path_to_output_file.xlsx'
# with pd.ExcelWriter(output_file_path) as writer:
#     combined_df.to_excel(writer, sheet_name='Combined_Sheet', index=False)





