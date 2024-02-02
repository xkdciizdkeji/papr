
import os
import re
import pandas as pd
import pandas as pd
import openpyxl
import numpy as np
from openpyxl import load_workbook
import glob

# import pathlib
# p = pathlib.Path('./')
# for txt_file in p.glob(pattern='*.txt'):
#     print (txt_file)


# 查找所有后缀为 .txt 的文本文件
txt_files = glob.glob('*.txt')
lines=[]
txt_file_names=[]
#print(txt_files)
for txt_file in txt_files:
# 读取文本文件并按行分割
    with open(txt_file, 'r',encoding="utf-8") as file:
        this_lines = file.readlines()
        lines=lines+this_lines
        txt_file_names.append(txt_file)

#print(lines)


# 定义正则表达式模式
case_pattern = re.compile(r'Simple_inputs/(.+)\.cap')

gcell_pattern = re.compile(r'gcell grid:\s+(\d+ x \d+ x \d+)')
unit_length_wire_pattern = re.compile(r'unit length wire:\s+(\d+\.\d+)')
unit_via_pattern = re.compile(r'unit via:\s+(\d+)')
num_nets_pattern = re.compile(r'num of nets :\s+(\d+)')
time_pattern = re.compile(r'\[\s*([\d.]+)\]')
step_routed_pattern = re.compile(r'step routed #nets:\s+(\d+),\s+(\d+),\s+(\d+)')
step_time_pattern = re.compile(r'step time consumption:\s+([\d.]+) s,\s+([\d.]+) s,\s+([\d.]+) s')
cost_pattern = re.compile(r'(wire|via|overflow|total) cost:\s+([\d.]+)')

num_of_open_nets_pattern = re.compile(r'use evaluator test data: Number of open nets :\s(\d+)')
evaluator_wirelength_cost_pattern= re.compile(r'use evaluator test data: wirelength cost\s+([\d.]+)')
evaluator_via_cost_pattern= re.compile(r'use evaluator test data: overflow cost\s+([\d.]+)')
evaluator_overflow_cost_pattern= re.compile(r'use evaluator test data: via cost\s+([\d.]+)')
evaluator_total_cost_pattern= re.compile(r'use evaluator test data: total cost\s+([\d.]+)')

cpp_write_txt_time_pattern = re.compile(r'stage 1: pattern routing:cpp_write_txt_time:\s+([\d.]+)')
call_python_script_time_pattern = re.compile(r'stage 1: pattern routing:call_python_script_total_time:\s+([\d.]+)')
python_script_read_txt_time_pattern = re.compile(r'python_script_read_txt_time:\s+([\d.]+)')
python_script_runtime_pattern = re.compile(r'python_script_runtime:\s+([\d.]+)')
python_script_write_txt_time_pattern = re.compile(r'python_script_write_txt_time:\s+([\d.]+)')
cpp_read_txt_time_pattern = re.compile(r'stage 1: pattern routing:cpp_read_txt_time:\s+([\d.]+)')



# 初始化数据存储字典
data = {
    'Environment': [],
    'score':[],
    'runtime_factor':[],
    'evaluator_total_cost': [],
    'Total_Cost': [],
    'Total_Time': [],
    'num of open nets': [],
    'Case_Name': [],
    'GCell_Grid': [],
    'Unit_Length_Wire': [],
    'Unit_Via': [],
    'Num_of_Nets': [],
    'Total_Time': [],
    'Parser_Time': [],
    'Parser_Time_Proportion': [],
    'Stage1_Routed_Nets_Num': [],
    'Stage1_Routed_Time': [],

    'Stage1_C++_write_txt_time': [],
    'Stage1_call_python_total_time': [],
    'Stage1_python_script_read_txt_time': [],
    'Stage1_python_script_runtime': [],
    'Stage1_python_script_write_txt_time': [],
    'Stage1_C++_read_txt_time': [],

    'Stage1_Routed_Time_Proportion': [],
    'Stage2_Routed_Nets_Num': [],
    'Stage2_Routed_Time': [],
    'Stage2_Routed_Time_Proportion': [],
    'Stage3_Routed_Nets_Num': [],
    'Stage3_Routed_Time': [],
    'Stage3_Routed_Time_Proportion': [],
    'evaluator_wirelength_cost': [],
    'Wire_Cost': [],
    'Wire_Cost_Proportion': [],
    'evaluator_via_cost': [],
    'Via_Cost': [],
    'Via_Cost_Proportion': [],
    'Overflow_Cost': [],
    'evaluator_overflow_cost': [],
    'Overflow_Cost_Proportion': []
}

# 逐行解析文本内容
case_name = ''
total_time = 0.0
parser_time = 0.0
wire_cost = 0.0
via_cost = 0.0
overflow_cost = 0.0


for line in lines:
    case_match = re.search(case_pattern, line)
    #print(case_match)
    gcell_match = re.search(gcell_pattern, line)
    unit_length_wire_match = re.search(unit_length_wire_pattern, line)
    unit_via_match = re.search(unit_via_pattern, line)
    num_nets_match = re.search(num_nets_pattern, line)
    time_match = re.search(time_pattern, line)
    step_routed_match = re.search(step_routed_pattern, line)
    #print(step_routed_match)
    step_time_match = re.search(step_time_pattern, line)
    cost_match = re.search(cost_pattern, line)

    num_of_open_nets_match = re.search(num_of_open_nets_pattern, line)
    #print(num_nets_match)
    evaluator_total_cost_match = re.search(evaluator_total_cost_pattern, line)
    evaluator_wirelength_cost_match = re.search(evaluator_wirelength_cost_pattern, line)
    evaluator_via_cost_match = re.search(evaluator_via_cost_pattern, line)
    evaluator_overflow_cost_match = re.search(evaluator_overflow_cost_pattern, line)

    cpp_write_txt_time_match = re.search(cpp_write_txt_time_pattern, line)
    #print(cpp_write_txt_time_match)
    call_python_time_match = re.search(call_python_script_time_pattern, line)
    #print(call_python_time_match)
    python_script_read_txt_time_match = re.search(python_script_read_txt_time_pattern, line)
    #print(python_script_read_txt_time_match)
    python_script_runtime_match = re.search(python_script_runtime_pattern, line)
    #print(python_script_runtime_match)
    python_script_write_txt_time_match = re.search(python_script_write_txt_time_pattern, line)
    #print(python_script_write_txt_time_match)
    cpp_read_txt_time_match = re.search(cpp_read_txt_time_pattern, line)
    #print(cpp_read_txt_time_match)

    if num_of_open_nets_match:
        data['num of open nets'].append(num_of_open_nets_match.group(1))
    if evaluator_total_cost_match:
        data['evaluator_total_cost'].append(evaluator_total_cost_match.group(1))
    if evaluator_wirelength_cost_match:
        data['evaluator_wirelength_cost'].append(evaluator_wirelength_cost_match.group(1))
    if evaluator_via_cost_match:
        data['evaluator_via_cost'].append(evaluator_via_cost_match.group(1))
    if evaluator_overflow_cost_match:
        data['evaluator_overflow_cost'].append(evaluator_overflow_cost_match.group(1))

    if cpp_write_txt_time_match:
        data['Stage1_C++_write_txt_time'].append(cpp_write_txt_time_match.group(1))
    if call_python_time_match:
        data['Stage1_call_python_total_time'].append(call_python_time_match.group(1))
    if python_script_read_txt_time_match:
        data['Stage1_python_script_read_txt_time'].append(python_script_read_txt_time_match.group(1))
    if python_script_runtime_match:
        data['Stage1_python_script_runtime'].append(python_script_runtime_match.group(1))
    if python_script_write_txt_time_match:
        data['Stage1_python_script_write_txt_time'].append(python_script_write_txt_time_match.group(1))
    if cpp_read_txt_time_match:
        data['Stage1_C++_read_txt_time'].append(cpp_read_txt_time_match.group(1))

    if case_match:
        #data['Environment'].append(txt_file)
        #data['Environment'].append(txt_file[ :txt_file.index('.')])
        case_name = case_match.group(1).split('/')[-1].split('.')[0]
        data['Case_Name'].append(case_name)
        #print(case_name)
    if gcell_match:
        data['GCell_Grid'].append(gcell_match.group(1))
        #print(gcell_match.group(1))
    if unit_length_wire_match:
        data['Unit_Length_Wire'].append(unit_length_wire_match.group(1))
        #print(unit_length_wire_match.group(1))
    if unit_via_match:
        data['Unit_Via'].append(unit_via_match.group(1))
        #print(unit_via_match.group(1))
    if num_nets_match:
        data['Num_of_Nets'].append(num_nets_match.group(1))
        #print(num_nets_match.group(1))
    if time_match:
        timestamp = float(time_match.group(1))
        #total_time = max(total_time, timestamp)
        #print(f"Total time: {total_time} seconds")
        # if 'total cost(ispd24 score)' in line:
        #     total_time = timestamp
        #     data['Total_Time'].append(total_time)
        #     #print(f"Total time: {total_time} seconds")
        if 'Finished parsing' in line:
            parser_time = timestamp
            data['Parser_Time'].append(parser_time)
            #print(f"Parser time: {parser_time} seconds")
        #if 'Terminated.' in line:
            
            
            #print(f"Parser time proportion: {data['Parser_Time_Proportion'][-1]}")
    if step_routed_match:
        data['Stage1_Routed_Nets_Num'].append(step_routed_match.group(1))
        #print(step_routed_match.group(1))
        data['Stage2_Routed_Nets_Num'].append(step_routed_match.group(2))
        #print(step_routed_match.group(2))
        data['Stage3_Routed_Nets_Num'].append(step_routed_match.group(3))
        #print(step_routed_match.group(3))

    if step_time_match:
        data['Stage1_Routed_Time'].append(step_time_match.group(1))
        
        
        
        #print(step_time_match.group(1))
        data['Stage2_Routed_Time'].append(step_time_match.group(2))
        #print(step_time_match.group(2))
        data['Stage3_Routed_Time'].append(step_time_match.group(3))
        #print(step_time_match.group(3))

        precise_total_time = float(step_time_match.group(1))+float(step_time_match.group(2))+float(step_time_match.group(3))+parser_time
        data['Total_Time'].append(precise_total_time)
        #print(precise_total_time)
        data['Stage1_Routed_Time_Proportion'].append(round(float(step_time_match.group(1)) / precise_total_time, 2))
        #print(f"Stage1 time proportion: {data['Stage1_Routed_Time_Proportion'][-1]}")
        data['Stage2_Routed_Time_Proportion'].append(round(float(step_time_match.group(2)) / precise_total_time, 2))
        #print(f"Stage2 time proportion: {data['Stage2_Routed_Time_Proportion'][-1]}")
        data['Stage3_Routed_Time_Proportion'].append(round(float(step_time_match.group(3)) / precise_total_time, 2))
        #print(f"Stage3 time proportion: {data['Stage3_Routed_Time_Proportion'][-1]}")
        data['Parser_Time_Proportion'].append(round(parser_time / precise_total_time, 2))
    if cost_match:
        cost_type = cost_match.group(1)
        #print(cost_match.group(1))
        #print(cost_type)
        cost_value = float(cost_match.group(2))
        if cost_type == 'wire':
            wire_cost = cost_value
            #print(f"Wire cost: {wire_cost}")
            data['Wire_Cost'].append(wire_cost)
        elif cost_type == 'via':
            via_cost = cost_value
            #print(f"Via cost: {via_cost}")
            data['Via_Cost'].append(via_cost)
        elif cost_type == 'overflow':
            overflow_cost = cost_value
            #print(f"Overflow cost: {overflow_cost}")
            data['Overflow_Cost'].append(overflow_cost)
            total_cost = wire_cost + via_cost + overflow_cost
            #print(f"Total cost: {total_cost}")
            data['Total_Cost'].append(total_cost)
            data['Wire_Cost_Proportion'].append(round((wire_cost / total_cost) * 100, 2))
            data['Via_Cost_Proportion'].append(round((via_cost / total_cost) * 100, 2))
            data['Overflow_Cost_Proportion'].append(round((overflow_cost / total_cost) * 100, 2))
print(data)
            
median_time = []
for txt_file_name in txt_file_names:
    for case_index in range(int(len(data['Case_Name'])/len(txt_file_names))):
        data['Environment'].append(txt_file_name[ :txt_file_name.index('.')])
        if data['Environment'][-1] == 'cpu':
            median_time.append(data['Total_Time'][len(data['Environment'])-1])

#print(median_time)

for txt_file_name in txt_file_names:
    for case_index in range(int(len(data['Case_Name'])/len(txt_file_names))):
        runtime_factor=[0.02*np.log2(float(data['Total_Time'][len(data['score'])])/float(median_time[case_index])),0.2,-0.2]
        #print(runtime_factor)
        data['runtime_factor'].append(str(np.median(runtime_factor)))
        data['score'].append(str(float(data['evaluator_total_cost'][len(data['score'])])*(np.median(runtime_factor)+1)))




print(data)

# 创建 DataFrame
df = pd.DataFrame(data)


if os.path.exists(os.path.join("",'output.xlsx')):#如果根目录下已经有了output.xlsx，则追加数据
    # 写入 Excel 文件
    with pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a') as writer:
        # 加载现有的 Excel 文件
        writer.workbook = load_workbook('output.xlsx')
        #print('writer.workbook')

        # 将数据写入现有的工作表
        for case in df['Case_Name'].unique():
            case_df = df[df['Case_Name'] == case]
            sheet_name = case

            # 如果工作表已存在，则追加数据
            if sheet_name in writer.workbook.sheetnames:
                
                startrow = writer.workbook[sheet_name].max_row
                for row in pd.DataFrame(case_df).iterrows():
                    values = row[1].to_list()
                    
                    for col in range(len(values)):
                        writer.workbook[sheet_name].cell(row=startrow + 1, column=col + 1, value=values[col])
                    
                    
            else:
                case_df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.workbook.save('output.xlsx')
    writer.workbook.close() 
else:#如果根目录下没有output.xlsx


    # 写入 Excel 文件
    with pd.ExcelWriter('output.xlsx') as writer:
        for case in df['Case_Name'].unique():
            case_df = df[df['Case_Name'] == case]
            case_df.to_excel(writer, sheet_name=case, index=False)
    #writer.workbook.save('output.xlsx')
    #writer.workbook.close()








from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 读取xlsx文件
file_path = 'output.xlsx'
workbook = load_workbook(file_path)

# 创建一个新的工作表用于拼接数据
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = 'Combined Data'

# 指定要处理的列
columns_to_process = [2,3,4,5,6,30,33,37] 

# 循环每个sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # 处理指定的列
    for col in columns_to_process:
        column_data = [sheet.cell(row=row, column=col).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None]
        if column_data:
            min_value = min(column_data)
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value == min_value:
                    sheet.cell(row=row, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 将每个sheet的内容拼接到新的工作表中
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        
        new_sheet.append([cell.value for cell in row])
        for cell in row:
            if cell.fill == PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"):
                #print(cell.fill)
                new_sheet.cell(row=new_sheet.max_row, column=cell.column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                new_sheet.cell(row=new_sheet.max_row, column=cell.column).font = openpyxl.styles.Font(bold=True, color="00003366")
        
        
    new_sheet.append([])
    # 调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 调整列宽
    for column in new_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        new_sheet.column_dimensions[column_letter].width = adjusted_width

# # 复制格式
# for row in new_sheet.iter_rows():
#     for cell in row:
#         if cell.fill is not None:
#             new_sheet.cell(row=cell.row, column=cell.column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#             #cell.fill = sheet.cell(cell.row, cell.column).fill

# 保存修改后的文件和新的工作表
workbook.save('经过数据标注的output.xlsx')
workbook.close
new_workbook.save('拼接在一个sheet里的经过数据标注的output.xlsx')














# # 读取原始 Excel 文件
# file_path = 'output.xlsx'
# xl = pd.ExcelFile(file_path)

# # 读取每个 sheet 的数据到 DataFrame
# dfs = {sheet_name: xl.parse(sheet_name) for sheet_name in xl.sheet_names}

# # 创建一个新的 DataFrame 用于存放所有 sheet 的数据
# combined_df = pd.DataFrame()

# # 将所有 sheet 的数据合并到一个 DataFrame 中，并在每个 DataFrame 之间插入一行空行
# for sheet_name, df in dfs.items():
#     if not combined_df.empty:
#         # 插入空行
#         empty_row = pd.Series([None] * len(combined_df.columns), index=combined_df.columns)
#         combined_df = pd.concat([combined_df, pd.DataFrame({col: [None] * 1})], ignore_index=True)
#     combined_df = pd.concat([combined_df, df], ignore_index=True)

# # 创建一个新的 Excel 文件并将合并后的数据写入新的 sheet
# output_file_path = 'path_to_output_file.xlsx'
# with pd.ExcelWriter(output_file_path) as writer:
#     combined_df.to_excel(writer, sheet_name='Combined_Sheet', index=False)





